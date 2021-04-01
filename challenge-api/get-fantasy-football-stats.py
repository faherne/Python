#!/usr/bin/python3
"""RZFeeser || exploring a catfact api
In this example, we query a fantasy football API source and output formatted data.
API Source:
https://fantasy.premierleague.com/api/bootstrap-static/
"""

import requests
from openpyxl import Workbook

API = 'https://fantasy.premierleague.com/api/bootstrap-static/'

def main():
    # Pull data from API
    r = requests.get(API)

    # determine if r is a "200"
    if r.status_code != 200:
        print(r.status_code)
        print(r.json())
        exit()

    # we want to grab the data off the 200 response
    ffdata = r.json()
   
    #Create workbook object
    wb = Workbook()
    wbsheet = wb.active
    wbsheet.title='FF Dream Team Stats'

    #Add titles in the first row of each column
    wbsheet.cell(row=1, column=1).value='Name'
    wbsheet.cell(row=1, column=2).value='Minutes'
    wbsheet.cell(row=1, column=3).value='GoalsScored'
    wbsheet.cell(row=1, column=4).value='Assists'
    wbsheet.cell(row=1, column=5).value='InDreamTeam'
    wbsheet.cell(row=1, column=6).value='PointsPerGame'

    for playername in ffdata['elements']:
        webname = playername.get("web_name")
        gameminutes = playername.get("minutes")
        goalsscored = playername.get("goals_scored")
        assists = playername.get("assists")
        indreamteam = playername.get("in_dreamteam")
        pointspergame = playername.get("points_per_game")
        
        if indreamteam:
            print(" Name: " + webname + ", Minutes: " + str(gameminutes))
            print(" Goals scored: " + str(goalsscored) + ", Assists: " + str(assists))
            print(" In Dreamteam: " + str(indreamteam) + ", Points per Game: " + pointspergame + "\n")
            
            #Loop to set the value of each cell
            for i in range(11):
                wbsheet.cell(row=i+2, column=1).value=webname
                wbsheet.cell(row=i+2, column=2).value=gameminutes
                wbsheet.cell(row=i+2, column=3).value=goalsscored
                wbsheet.cell(row=i+2, column=4).value=assists
                wbsheet.cell(row=i+2, column=5).value=indreamteam
                wbsheet.cell(row=i+2, column=6).value=pointspergame

    #Finally, save the file and give it a name
    wb.save('~\mycode\challenge-api\FF_Dream_Team_Stats.xlsx')

if __name__ == "__main__":
    main()
